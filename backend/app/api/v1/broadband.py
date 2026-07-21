from datetime import datetime, date
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, col
from app.core.database import get_session
from app.models.broadband import (
    BroadbandContract, BroadbandContractCreate,
    BroadbandContractUpdate, BroadbandContractRead,
    RENEWAL_CYCLE_MONTHS, RENEWAL_CYCLE_LABELS, calc_annual_cost,
)
from app.services.dingtalk import send_renewal_reminder, send_test_message
from app.services.broadband_renewal import get_next_renewal
import io
import openpyxl
from openpyxl.styles import Font, PatternFill

router = APIRouter()


@router.get('/dashboard', summary='宽带合同仪表盘')
def broadband_dashboard(session: Session = Depends(get_session)):
    today = date.today()
    all_contracts = session.exec(select(BroadbandContract)).all()
    active = [c for c in all_contracts if c.status == 'active']
    expired = [c for c in all_contracts if c.status == 'expired']
    expiring_30 = [c for c in active if (c.contract_end - today).days <= 30 and (c.contract_end - today).days >= 0]
    expiring_7 = [c for c in active if (c.contract_end - today).days <= 7 and (c.contract_end - today).days >= 0]
    # 按续费周期对齐的即将到期统计
    expiring_renewal_30 = 0
    expiring_renewal_7 = 0
    for c in active:
        dr = get_next_renewal(c, today)['days_remaining']
        if 0 <= dr <= 30:
            expiring_renewal_30 += 1
        if 0 <= dr <= 7:
            expiring_renewal_7 += 1
    total_annual = sum(c.annual_cost or 0 for c in active)
    return {
        'total': len(all_contracts),
        'active': len(active),
        'expired': len(expired),
        'expiring_30d': len(expiring_30),
        'expiring_7d': len(expiring_7),
        'expiring_renewal_30d': expiring_renewal_30,
        'expiring_renewal_7d': expiring_renewal_7,
        'total_annual_cost': round(total_annual, 2),
    }


@router.get('', response_model=dict, summary='分页查询宽带合同')
def list_contracts(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    session: Session = Depends(get_session),
):
    q = select(BroadbandContract)
    if keyword:
        q = q.where(
            col(BroadbandContract.provider).contains(keyword)
            | col(BroadbandContract.circuit_id).contains(keyword)
            | col(BroadbandContract.location).contains(keyword)
            | col(BroadbandContract.contact_name).contains(keyword)
        )
    if status:
        q = q.where(BroadbandContract.status == status)
    total = len(session.exec(q).all())
    items = session.exec(q.offset((page - 1) * size).limit(size)).all()
    today = date.today()
    result_items = []
    for c in items:
        item = c.model_dump(mode='json')
        renewal = get_next_renewal(c, today)
        item['next_renewal_deadline'] = renewal['next_deadline'].isoformat()
        item['next_renewal_days'] = renewal['days_remaining']
        item['deadline_type'] = renewal['deadline_type']
        result_items.append(item)
    return {'total': total, 'page': page, 'size': size, 'items': result_items}


@router.post('', response_model=BroadbandContractRead, status_code=201)
def create_contract(data: BroadbandContractCreate, session: Session = Depends(get_session)):
    # 如果 renewal_cost 有值但 annual_cost 没填，自动计算年费
    if data.renewal_cost is not None and data.annual_cost is None:
        data.annual_cost = calc_annual_cost(data.renewal_cost, data.renewal_cycle or 'annual')
    contract = BroadbandContract.model_validate(data)
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.get('/{contract_id}', response_model=dict)
def get_contract(contract_id: int, session: Session = Depends(get_session)):
    contract = session.get(BroadbandContract, contract_id)
    if not contract:
        raise HTTPException(404, '合同不存在')
    today = date.today()
    item = contract.model_dump(mode='json')
    renewal = get_next_renewal(contract, today)
    item['next_renewal_deadline'] = renewal['next_deadline'].isoformat()
    item['next_renewal_days'] = renewal['days_remaining']
    item['deadline_type'] = renewal['deadline_type']
    return item


@router.put('/{contract_id}', response_model=BroadbandContractRead)
def update_contract(contract_id: int, data: BroadbandContractUpdate, session: Session = Depends(get_session)):
    contract = session.get(BroadbandContract, contract_id)
    if not contract:
        raise HTTPException(404, '合同不存在')
    update_data = data.model_dump(exclude_unset=True)
    
    # 如果更新了 renewal_cost 或 renewal_cycle 但没更新 annual_cost，自动计算
    cost_changed = 'renewal_cost' in update_data or 'renewal_cycle' in update_data
    if cost_changed and 'annual_cost' not in update_data:
        new_cost = update_data.get('renewal_cost', contract.renewal_cost)
        new_cycle = update_data.get('renewal_cycle', contract.renewal_cycle)
        if new_cost is not None:
            update_data['annual_cost'] = calc_annual_cost(new_cost, new_cycle)
    
    for k, v in update_data.items():
        setattr(contract, k, v)
    contract.updated_at = datetime.utcnow()
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.delete('/{contract_id}', status_code=204)
def delete_contract(contract_id: int, session: Session = Depends(get_session)):
    contract = session.get(BroadbandContract, contract_id)
    if not contract:
        raise HTTPException(404, '合同不存在')
    session.delete(contract)
    session.commit()


@router.post('/{contract_id}/test-notify', summary='发送测试钉钉通知')
def test_notify(contract_id: int, session: Session = Depends(get_session)):
    contract = session.get(BroadbandContract, contract_id)
    if not contract:
        raise HTTPException(404, '合同不存在')
    renewal = get_next_renewal(contract)
    ok = send_renewal_reminder(
        provider=contract.provider,
        circuit_id=contract.circuit_id,
        bandwidth_mbps=contract.bandwidth_mbps,
        location=contract.location,
        contract_end=contract.contract_end,
        renewal_deadline=renewal['next_deadline'],
        days_remaining=renewal['days_remaining'],
        contact_name=contract.contact_name,
        deadline_type=renewal['deadline_type'],
        renewal_cycle=contract.renewal_cycle,
        renewal_cost=contract.renewal_cost,
        annual_cost=contract.annual_cost,
    )
    if ok:
        return {'success': True, 'message': '通知已发送'}
    else:
        raise HTTPException(500, '通知发送失败，请检查钉钉 Webhook 配置')


# ── Excel 导入导出 ──────────────────────────────

@router.get('/export/template', summary='下载宽带导入模板')
def download_template():
    """下载 Excel 导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '宽带合同导入模板'

    headers = ['运营商*', '带宽(Mbps)*', '合同开始*', '合同到期*',
               '续费周期', '周期费用(元)', '年费(元)', '线路编号',
               '位置', '联系人', '联系电话', '月费(元)',
               '自动续费(是/否)', '状态', '备注', '提醒天数']
    
    fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    font = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill

    # 示例数据
    sample = ['中国电信', '100', '2025-01-01', '2025-12-31',
              'quarterly', '3000', '12000', 'CHN-001',
              '数据中心A', '张三', '13800138000', '1000',
              '是', 'active', '主用线路', '30,15,7']
    for col_idx, v in enumerate(sample, 1):
        ws.cell(row=2, column=col_idx, value=v)

    # 设置列宽
    widths = [14, 14, 14, 14, 14, 14, 14, 14, 16, 10, 14, 12, 16, 10, 20, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=broadband_import_template.xlsx'},
    )


@router.post('/import/excel', summary='从 Excel 批量导入宽带合同')
def import_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    """上传 Excel 文件，批量创建宽带合同"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, '请上传 .xlsx 格式的 Excel 文件')

    contents = file.file.read()
    buf = io.BytesIO(contents)
    try:
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
    except Exception:
        raise HTTPException(400, '无法解析 Excel 文件，请确认格式正确')

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows or all(c is None for c in rows[0]):
        raise HTTPException(400, 'Excel 文件为空，请先填写数据')

    # 字段映射（按模板列顺序）
    imported = 0
    errors = []
    cycle_map = {'每月': 'monthly', '月': 'monthly', '每季度': 'quarterly', '季度': 'quarterly',
                 '每半年': 'semi_annual', '半年': 'semi_annual', '每年': 'annual', '年': 'annual'}

    def _parse_date(val):
        """解析日期：支持 datetime 对象和字符串"""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if hasattr(val, 'date'):  # date 对象
            return val
        s = str(val).strip()
        for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y/%m/%d %H:%M:%S', '%m/%d/%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise ValueError(f'无法解析日期: {s}')

    def _parse_num(val):
        """解析数字：支持 int/float/str"""
        if val is None:
            return None
        return float(val)

    for idx, row in enumerate(rows, 2):
        try:
            # 跳过列数不足或运营商为空
            if len(row) < 4 or not row[0]:
                if any(c is not None for c in row[:4]):
                    errors.append(f'第{idx}行：数据不完整，跳过')
                continue

            provider = str(row[0] or '').strip()
            if not provider:
                continue  # 跳过空行

            bandwidth_mbps = int(_parse_num(row[1])) if row[1] is not None else None
            contract_start = _parse_date(row[2])
            contract_end = _parse_date(row[3])

            if not all([bandwidth_mbps, contract_start, contract_end]):
                errors.append(f'第{idx}行：缺少必填字段（带宽、合同开始、合同到期）')
                continue

            # 解析续费周期
            raw_cycle = str(row[4] or '').strip() if len(row) > 4 else ''
            renewal_cycle = cycle_map.get(raw_cycle, raw_cycle if raw_cycle in RENEWAL_CYCLE_MONTHS else 'annual')
            # 解析费用
            renewal_cost = _parse_num(row[5]) if len(row) > 5 else None
            annual_cost = _parse_num(row[6]) if len(row) > 6 else None
            circuit_id = (str(row[7] or '').strip() or None) if len(row) > 7 else None
            location = (str(row[8] or '').strip() or None) if len(row) > 8 else None
            contact_name = (str(row[9] or '').strip() or None) if len(row) > 9 else None
            contact_phone = (str(row[10] or '').strip() or None) if len(row) > 10 else None
            monthly_cost = _parse_num(row[11]) if len(row) > 11 else None
            auto_renew = str(row[12] or '').strip() in ('是', 'Y', 'yes', 'true', '1') if len(row) > 12 else False
            status = str(row[13] or '').strip() if len(row) > 13 else 'active'
            if status not in ('active', 'expired', 'cancelled'):
                status = 'active'
            notes = (str(row[14] or '').strip() or None) if len(row) > 14 else None
            reminder_days = str(row[15] or '').strip() if (len(row) > 15 and row[15]) else '30,15,7'

            # 自动计算年费
            if annual_cost is None:
                annual_cost = calc_annual_cost(renewal_cost, renewal_cycle)

            contract = BroadbandContract(
                provider=provider,
                circuit_id=circuit_id,
                bandwidth_mbps=bandwidth_mbps,
                location=location,
                renewal_cycle=renewal_cycle,
                renewal_cost=renewal_cost,
                annual_cost=annual_cost,
                monthly_cost=monthly_cost,
                contract_start=contract_start,
                contract_end=contract_end,
                auto_renew=auto_renew,
                contact_name=contact_name,
                contact_phone=contact_phone,
                reminder_days=reminder_days,
                status=status,
                notes=notes,
            )
            session.add(contract)
            imported += 1
        except Exception as e:
            errors.append(f'第{idx}行：{str(e)}')

    if imported > 0:
        session.commit()

    return {
        'imported': imported,
        'errors': errors,
        'message': f'成功导入 {imported} 条记录{f"，{len(errors)} 条错误" if errors else ""}',
    }
